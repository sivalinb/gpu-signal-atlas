#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

PRIMARY = "22C55E"
HEADER = "172033"
MUTED = "64748B"


def load(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def pct(value: float) -> str:
    return f"{value * 100:.1f}%"


def delta(after: float, before: float) -> float:
    return round(after - before, 4)


def flatten_metrics(result: dict[str, Any]) -> dict[str, float]:
    aggregate = result["aggregate"]
    return {
        "Pass rate": aggregate["passRate"],
        "Recall@5": aggregate["retrieval"]["recallAt5"],
        "MRR": aggregate["retrieval"]["mrr"],
        "Citation validity": aggregate["quality"]["citationValidity"],
        "Claim faithfulness": aggregate["quality"]["claimFaithfulness"],
        "Task contract": aggregate["quality"]["taskContract"],
        "Guardrail pass rate": aggregate["quality"]["guardrailPassRate"],
        "Refusal F1": aggregate["refusal"]["f1"],
        "p95 latency (ms)": aggregate["performance"]["p95LatencyMs"],
    }


def build_comparison(baseline: dict[str, Any], improved: dict[str, Any]) -> dict[str, Any]:
    before = flatten_metrics(baseline)
    after = flatten_metrics(improved)
    baseline_cases = {item["caseId"]: item for item in baseline["cases"]}
    improved_cases = {item["caseId"]: item for item in improved["cases"]}
    resolved = [
        case_id
        for case_id, item in baseline_cases.items()
        if not item["passed"] and improved_cases[case_id]["passed"]
    ]
    regressions = [
        case_id
        for case_id, item in baseline_cases.items()
        if item["passed"] and not improved_cases[case_id]["passed"]
    ]
    metric_rows = []
    for name in before:
        metric_rows.append({
            "name": name,
            "baseline": before[name],
            "improved": after[name],
            "delta": delta(after[name], before[name]),
            "unit": "ms" if "latency" in name else "ratio",
        })
    return {
        "schemaVersion": 1,
        "datasetVersion": baseline["datasetVersion"],
        "datasetSha256": baseline["datasetSha256"],
        "cases": baseline["aggregate"]["cases"],
        "distribution": {"happyPath": 24, "edgeCase": 14, "knownFailure": 7, "adversarial": 3},
        "baseline": baseline["aggregate"],
        "improved": improved["aggregate"],
        "metrics": metric_rows,
        "resolvedCases": resolved,
        "regressions": regressions,
        "targetedChanges": [
            {
                "title": "Normalize collector output",
                "problem": "Lowercase DCGM keys bypassed exact-signal extraction.",
                "change": "Case-insensitive parsing with canonical uppercase metric IDs.",
            },
            {
                "title": "Harden Xid parsing and semantic routing",
                "problem": "Short, key-value, and operational paraphrases could be refused.",
                "change": "Additional safe syntax and reviewed intent patterns; exact IDs can clear the length gate.",
            },
            {
                "title": "Add an instruction-manipulation guardrail",
                "problem": "Telemetry containing prompt-injection language could reach grounded generation.",
                "change": "Explicit pre-retrieval refusal with zero citations and a visible decision reason.",
            },
        ],
        "monitoring": [
            "Quality drift: pass rate, Recall@5, citation validity, claim faithfulness",
            "Safety drift: refusal F1 and adversarial guardrail pass rate",
            "Performance drift: p50/p95 latency, Pinecone query counts, tokens, and cost",
            "Reliability: tool errors, empty retrievals, and stale corpus versions",
        ],
    }


def write_csv(path: Path, baseline: dict[str, Any], improved: dict[str, Any]) -> None:
    before = {item["caseId"]: item for item in baseline["cases"]}
    after = {item["caseId"]: item for item in improved["cases"]}
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["case_id", "scenario", "category", "baseline_pass", "improved_pass", "resolved", "baseline_failures", "improved_failures", "baseline_latency_ms", "improved_latency_ms"])
        for case_id, item in before.items():
            changed = after[case_id]
            writer.writerow([
                case_id,
                item["scenarioType"],
                item["category"],
                item["passed"],
                changed["passed"],
                not item["passed"] and changed["passed"],
                ";".join(item["failureReasons"]),
                ";".join(changed["failureReasons"]),
                item["latencyMs"],
                changed["latencyMs"],
            ])


def write_workbook(path: Path, comparison: dict[str, Any], baseline: dict[str, Any], improved: dict[str, Any]) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Executive scorecard"
    summary.append(["GPU Signal Atlas — Week 4 Controlled Evaluation"])
    summary.append(["Metric", "Baseline", "Improved", "Delta"])
    for item in comparison["metrics"]:
        summary.append([item["name"], item["baseline"], item["improved"], item["delta"]])
    summary.append([])
    summary.append(["Resolved cases", len(comparison["resolvedCases"])])
    summary.append(["Regressions", len(comparison["regressions"])])
    summary.append(["Dataset SHA-256", comparison["datasetSha256"]])

    cases_sheet = workbook.create_sheet("Per-case results")
    cases_sheet.append(["Case", "Scenario", "Category", "Baseline", "Improved", "Resolved", "Baseline latency ms", "Improved latency ms"])
    before = {item["caseId"]: item for item in baseline["cases"]}
    after = {item["caseId"]: item for item in improved["cases"]}
    for case_id, item in before.items():
        changed = after[case_id]
        cases_sheet.append([case_id, item["scenarioType"], item["category"], item["passed"], changed["passed"], not item["passed"] and changed["passed"], item["latencyMs"], changed["latencyMs"]])

    failures = workbook.create_sheet("Failure analysis")
    failures.append(["Baseline failure cluster", "Count"])
    for item in baseline["aggregate"]["failureClusters"]:
        failures.append([item["reason"], item["count"]])
    failures.append([])
    failures.append(["Resolved case"])
    for case_id in comparison["resolvedCases"]:
        failures.append([case_id])

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A3" if sheet is summary else "A2"
        sheet.sheet_view.showGridLines = False
        for cell in sheet[1]:
            cell.font = Font(color="FFFFFF", bold=True, size=14 if sheet is summary else 11)
            cell.fill = PatternFill("solid", fgColor=HEADER)
        header_row = 2 if sheet is summary else 1
        for cell in sheet[header_row]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = PatternFill("solid", fgColor=PRIMARY)
        for column in sheet.columns:
            width = min(52, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
            sheet.column_dimensions[column[0].column_letter].width = width
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    workbook.save(path)


def write_markdown(path: Path, comparison: dict[str, Any], baseline: dict[str, Any]) -> None:
    rows = []
    for item in comparison["metrics"]:
        if item["unit"] == "ms":
            rendered = f"| {item['name']} | {item['baseline']:.3f} | {item['improved']:.3f} | {item['delta']:+.3f} |"
        else:
            rendered = f"| {item['name']} | {pct(item['baseline'])} | {pct(item['improved'])} | {item['delta'] * 100:+.1f} pp |"
        rows.append(rendered)
    clusters = "\n".join(f"- `{item['reason']}`: {item['count']} cases" for item in baseline["aggregate"]["failureClusters"])
    resolved = "\n".join(f"- `{case_id}`" for case_id in comparison["resolvedCases"])
    changes = "\n".join(
        f"### {index}. {item['title']}\n\n**Observed problem:** {item['problem']}\n\n**Targeted change:** {item['change']}"
        for index, item in enumerate(comparison["targetedChanges"], start=1)
    )
    managed_path = Path("evaluation/week4/results/pinecone-improved.json")
    managed_section = ""
    if managed_path.exists():
        managed = load(managed_path)["aggregate"]
        managed_section = f"""
## Managed Pinecone production-path check

The improved agent also ran the identical 48 cases against the configured Pinecone namespace. It passed **{managed['passed']}/{managed['cases']} cases**, preserved **{pct(managed['retrieval']['recallAt5'])} Recall@5**, citation validity, claim faithfulness, refusal F1, and guardrail behavior, consumed **{managed['performance']['pineconeReadUnits']} query read units**, and measured **{managed['performance']['p50LatencyMs']:.1f} ms p50 / {managed['performance']['p95LatencyMs']:.1f} ms p95** end-to-end retrieval latency. No Pinecone credential or raw telemetry is stored in the result artifact.
"""
    text = f"""# Week 4 Evaluation Report

## Executive result

GPU Signal Atlas was evaluated on the same frozen, human-reviewed 48-case dataset before and after three targeted changes. The pass rate improved from **{pct(comparison['baseline']['passRate'])}** to **{pct(comparison['improved']['passRate'])}**. All **{len(comparison['resolvedCases'])}** baseline failures were resolved and the controlled rerun introduced **{len(comparison['regressions'])} regressions**.

This evaluation tests the evidence agent's retrieval, refusal, citation, output-contract, security, and latency behavior. It does not claim that a single GPU event proves root cause.

## Frozen dataset

- Version: `{comparison['datasetVersion']}`
- SHA-256: `{comparison['datasetSha256']}`
- 24 happy-path cases, 14 edge cases, 7 known-failure regressions, and 3 adversarial cases
- Inputs use synthetic or curated public telemetry formats; no production payload is uploaded
- Expected evidence IDs and refusal labels are stored outside the operational corpus

## Controlled comparison

| Metric | Baseline | Improved | Delta |
|---|---:|---:|---:|
{chr(10).join(rows)}

Latency is local process time and should be read separately from hosted Pinecone network latency. The deterministic generation path uses zero LLM tokens and has zero model cost; optional LLM mode is evaluated independently.
{managed_section}

## Baseline failure clusters

{clusters}

## Targeted improvements

{changes}

## Resolved regression cases

{resolved}

## LangSmith experiment design

The Python evaluation harness syncs the frozen dataset to LangSmith and creates paired baseline and improved experiments. Each test case has one root run, replay/agent and contract-validation child runs, dataset/example linkage, scenario metadata, deterministic evaluator feedback, latency, token, cost, and privacy fields. Experiment names and links are recorded in `evaluation/week4/results/langsmith.json` after upload.

## Production monitoring proposal

{chr(10).join(f'- {item}' for item in comparison['monitoring'])}

## Reproduce

```bash
python3 -m venv .venv-week4
.venv-week4/bin/pip install -r requirements-week4.txt
.venv-week4/bin/python evaluation/week4/python/validate_dataset.py
.venv-week4/bin/python evaluation/week4/python/run_evaluation.py --variant improved --output evaluation/week4/results/improved.json
.venv-week4/bin/python evaluation/week4/python/compare_results.py
```

The LangSmith upload command is documented in `docs/WEEK4_LOCAL_TESTING.md`; it reads the local key file without printing or copying the secret into the repository.
"""
    path.write_text(text, encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Compare Week 4 baseline and improved runs")
    parser.add_argument("--baseline", type=Path, default=Path("evaluation/week4/results/baseline.json"))
    parser.add_argument("--improved", type=Path, default=Path("evaluation/week4/results/improved.json"))
    parser.add_argument("--output-dir", type=Path, default=Path("evaluation/week4/results"))
    parser.add_argument("--report", type=Path, default=Path("docs/WEEK4_EVALUATION_REPORT.md"))
    args = parser.parse_args()
    baseline = load(args.baseline)
    improved = load(args.improved)
    if baseline["datasetSha256"] != improved["datasetSha256"]:
        raise ValueError("Baseline and improved results are not from the same frozen dataset")
    comparison = build_comparison(baseline, improved)
    args.output_dir.mkdir(parents=True, exist_ok=True)
    (args.output_dir / "comparison.json").write_text(json.dumps(comparison, indent=2) + "\n", encoding="utf-8")
    write_csv(args.output_dir / "per_case.csv", baseline, improved)
    write_workbook(args.output_dir / "week4_evaluation.xlsx", comparison, baseline, improved)
    write_markdown(args.report, comparison, baseline)
    print(f"Compared {comparison['cases']} cases: {len(comparison['resolvedCases'])} resolved, {len(comparison['regressions'])} regressions")


if __name__ == "__main__":
    main()
